# DB
import cx_Oracle
import os
import pymysql.cursors

# Excel
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side

import traceback
from typing import List, Dict, Optional, Literal, Any
from enum import Enum
from dataclasses import dataclass
from datetime import datetime


class Constant:
    # oracle client 경로
    LOCATION = r"C:\\...\\instantclient_23_6"


class DbConfig:
    """
    DB 연결정보
    """

    HOST = "192.168.xxx.xxx"
    PORT = 1234
    USER = "user"
    PASSWORD = "password"
    DATABASE = "db"
    CHARSET = "utf8mb4"  # MySQL

class Dbms(Enum):
    """
    DBMS 종류
    """

    ORACLE = "ORACLE"
    MYSQL = "MYSQL"


# ======== oracle client 경로 설정 ========
os.environ["PATH"] = Constant.LOCATION + ";" + os.environ["PATH"]

# ======== DB 정보 DTO ========
@dataclass
class TableInfo:
    """
    테이블 정보

    Attributes:
        owner (str): 소유자, user
        table_name (str): 테이블명
        tablespace_name (Optional[str]): 테이블스페이스
        comment (Optional[str]): 테이블 설명명
    """

    owner: str
    table_name: str
    tablespace_name: Optional[str]
    comment: Optional[str]

    def __init__(self, table_info: Dict[str, Any], dbms: Dbms):
        if dbms == Dbms.ORACLE:
            self.owner = table_info["OWNER"]
            self.table_name = table_info["TABLE_NAME"]
            self.tablespace_name = table_info["TABLESPACE_NAME"]
            self.comment = table_info["COMMENTS"]
        elif dbms == Dbms.MYSQL:
            self.owner = table_info["TABLE_SCHEMA"].upper()
            self.table_name = table_info["TABLE_NAME"].upper()
            self.tablespace_name = None
            self.comment = table_info["TABLE_COMMENT"]


@dataclass
class ColumnInfo:
    """
    컬럼 정보

    Attributes:
        column_name (str): 컬럼명
        data_type (str): 컬럼타입
        data_length (Optional[int]): 데이터 길이
        nullable (Literal["Y", "N"]): NULL 허용 여부
        key_type (Optional[str]): PK, FK, IDX
        data_default (Optional[str]): 기본값
        comment (Optional[str]): 설명
    """

    column_name: str
    data_type: str
    data_length: Optional[int]
    nullable: Literal["Y", "N"]
    key_type: Optional[str]
    data_default: Optional[str]
    comment: Optional[str]

    def __init__(self, row: Dict[str, Any], dbms: Dbms):
        if dbms == Dbms.ORACLE:
            self.column_name = row["COLUMN_NAME"]
            self.data_type = row["DATA_TYPE"]
            self.data_length = row["DATA_LENGTH"]
            self.nullable = row["NULLABLE"]
            self.key_type = row["KEY_TYPE"]
            self.data_default = row["DATA_DEFAULT"]
            self.comment = row["COL_COMMENTS"]
        elif dbms == Dbms.MYSQL:
            self.column_name = row["Field"].upper()
            self.data_type = row["Type"].upper()
            self.data_length = None
            self.nullable = "Y" if row["Null"] == "YES" else "N"
            self.key_type = row["Key"]
            self.data_default = row["Default"]
            self.comment = row["Comment"]


def main():
    """
    진입함수
    """

    schema = input("스키마 입력 > ")
    file_name = input("저장할 파일명 > ")

    format = "%Y-%m-%d %H:%M:%S"

    print(f"----------- [{datetime.now().strftime(format)}] START TO SAVE EXCEL FILE")

    try:
        conn = None

        # 1. DB 연결
        # conn = cx_Oracle.connect(
        #     DbConfig.USER,
        #     DbConfig.PASSWORD,
        #     f"{DbConfig.HOST}:{DbConfig.PORT}/{DbConfig.DATABASE}",
        # )
        conn = pymysql.connect(
            host=DbConfig.HOST,
            port=DbConfig.PORT,
            user=DbConfig.USER,
            password=DbConfig.PASSWORD,
            database=DbConfig.DATABASE,
            charset=DbConfig.CHARSET,
            cursorclass=pymysql.cursors.DictCursor,
        )

        with conn.cursor() as cursor:
            # 2. 실행
            # exec_oracle(cursor=cursor, schema=schema, file_name=file_name)
            exec_mysql(cursor=cursor, schema=schema, file_name=file_name)

    except Exception as e:
        print(f"## FAIL TO SAVE FILE :: {e}")
        traceback.print_exc()
    else:
        print(f"💚 SUCCESS TO SAVE FILE!")

    print(f"----------- [{datetime.now().strftime(format)}] FINISH TO SAVE EXCEL FILE")


def exec_oracle(
        cursor,
        schema: str,
        file_name: str,
):
    """
    Oracle 실행
    """

    # 엑셀 생성
    wb = Workbook()
    sheet = wb.active

    # 테이블 리스트 조회
    sql = f"""
        SELECT * 
        FROM ALL_TABLES AT
        LEFT JOIN ALL_TAB_COMMENTS ATC
        ON AT.TABLE_NAME = ATC.TABLE_NAME
        AND AT.OWNER = '{schema}'
        """

    cursor.execute(sql)
    tables = cursor.fetchall()

    for t in tables:
        t = t["TABLE_NAME"]

        # 테이블 정보 조회
        sql = f"""
            SELECT
                AT.TABLE_NAME
                , AT.TABLESPACE_NAME
                , AT.OWNER
                , ATC.COMMENTS
            FROM ALL_TABLES AT
            LEFT JOIN ALL_TAB_COMMENTS ATC
            ON AT.TABLE_NAME = ATC.TABLE_NAME
            WHERE AT.TABLE_NAEM = '{t}'
            AND AT.OWNER = '{schema}'
            """
        cursor.execute(sql)
        table_info = cursor.fetchone()

        # 컬럼정보 조회
        sql = f"""
            SELECT *
            FROM ALL_TAB_COLUMNS
            WHERE TABLE_NAME = '{t}'
            """
        cursor.execute(sql)
        column_list = cursor.fetchall()

        write_excel(
            sheet=sheet,
            table_info=TableInfo(table_info, Dbms.ORACLE),
            column_list=column_list,
            dbms=Dbms.ORACLE,
        )

    wb.save(f"{file_name}.xlsx")


def exec_mysql(
        cursor: pymysql.cursors.DictCursor,
        schema: str,
        file_name: str,
):
    """
    MySQL 실행
    """

    # 엑셀 생성
    wb = Workbook()
    sheet = wb.active

    # 테이블 리스트 조회
    sql = f"""
        SELECT TABLE_NAME
        FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_SCHEMA = '{schema}'
        """

    cursor.execute(sql)
    tables = cursor.fetchall()

    for t in tables:
        t = t["TABLE_NAME"]

        # 테이블 정보 조회
        sql = f"""
            SELECT 
                TABLE_NAME
                , TABLE_SCHEMA
                , TABLE_COMMENT
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_NAME = '{t}'
            """
        cursor.execute(sql)
        table_info = cursor.fetchone()

        # 컬럼정보 조회
        sql = f"""
            SHOW FULL COLUMNS FROM {t.upper()}
            """
        cursor.execute(sql)
        column_list = cursor.fetchall()

        write_excel(
            sheet=sheet,
            table_info=TableInfo(table_info, Dbms.MYSQL),
            column_list=column_list,
            dbms=Dbms.MYSQL,
        )

    wb.save(f"{file_name}.xlsx")


# ============= Excel 관련 START =============
def write_excel(sheet: Worksheet, table_info, column_list: List, dbms: Dbms):
    """
    엑셀에 데이터 입력
    """

    # 제목
    sheet.append(["테이블정의서"])

    # 테이블정보
    sheet.append(
        [
            "테이블명",
            table_info.table_name,
            "테이블스키마",
            table_info.owner,
        ]
    )

    # 컬럼 정보
    sheet.append(["번호", "컬럼명", "설명", "타입", "길이", "NULL", "KEY", "기본값"])

    # 컬럼 정보는 배열이므로 루프에서 실행합니다.
    i = 0
    for c in column_list:
        i += 1
        # DTO 변환
        column_info = ColumnInfo(c, dbms)

        sheet.append(
            [
                i,
                column_info.column_name,
                column_info.comment,
                column_info.data_type,
                column_info.data_length,
                column_info.nullable,
                column_info.key_type,
                column_info.data_default,
            ]
        )


main()
