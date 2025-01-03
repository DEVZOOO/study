"""
DB 명세서 엑셀 자동화
"""

# DB
import pymysql.cursors

import cx_Oracle
import os


# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side

# gui
from tkinter import *

from enum import Enum
from dataclasses import dataclass
from typing import Optional, Literal, List
import traceback
from datetime import datetime


class Constant:
    # oracle client 경로
    LOCATION = r"C:\\..."


class MySqlConfig:
    """
    MySQL 연결정보

    Attributes:
        HOST (str): url
        PORT (int): 연결포트
        USER (str): 아이디
        PASSWORD (str): 비밀번호
        DATABASE (str): DB명
        CHARSET (str): 인코딩

    """

    HOST = "xxx.xxx.xxx.xxx"
    PORT = 1234
    USER = "user"
    PASSWORD = "password"
    DATABASE = "db"
    CHARSET = "utf8mb4"


class OracleConfig:
    """
    Oracle 연결정보

    Attributes:
        HOST (str): url
        PORT (int): 연결포트트
        USER (str): 아이디
        PASSWORD (str): 비밀번호
        DATABASE (str): DB명
    """

    HOST = "xxx.xxx.xxx.xxx"
    PORT = 1234
    USER = "user"
    PASSWORD = "password"
    DATABASE = "db"


# ======== tkinter 관련 객체 ========
class TkObj:
    """
    tkinter 객체
    """

    TK_ROOT = None  # tk root
    # Frame
    TK_FRAME_FORM = None  # form frame
    TK_FRAME_RESULT = None  # result frame
    # Input
    TK_RADIO_DBMS = None  # DBMS 종류
    TK_ENTRY_SCHEMA = None  # 스키마명
    TK_RADIO_TABLE_LIST = None  # 테이블 리스트 조회 방법 - 전체 ALL / 특정 경로 PATH
    TK_ENTRY_TABLE_PATH = None  # 테이블 리스트 파일 경로
    TK_RADIO_SHEET = None  # 시트 분리여부 - 한 시트 ONE / 시트 분리 MULTI
    TK_RADIO_OUTPUT_TYPE = None  # 출력 형식 - 스타일 적용 STYLED / 기본 RAW
    TK_ENTRY_OUTPUT_PATH = None  # 출력 경로


class TkSize:
    """
    GUI 사이즈
    """

    WIDTH = 400
    HEIGHT = 300
    GAP = 20


class RadioDbms(Enum):
    """
    DBMS 종류 옵션
    - MYSQL
    - ORACLE
    """

    MYSQL = "MYSQL"
    ORACLE = "ORACLE"


class RadioTableList(Enum):
    """
    테이블 리스트 조회 방법 옵션
    - ALL
    - PATH
    """

    ALL = "전체"
    PATH = "특정 경로로"


class RadioSheet(Enum):
    """
    시트 분리여부 옵션
    - ONE
    - MULTI
    """

    ONE = "한 시트"
    MULTI = "시트 분리"


class RadioOutputType(Enum):
    """
    출력 형식 옵션

    - STYLED
    - RAW
    """

    STYLED = "스타일 적용"
    RAW = "기본"


class ExcelStyle:
    """
    엑셀 named style명
    - COMMON: 공통 적용
    - HIGHLIGHT: 강조
    - TITLE: 제목
    """

    COMMON = "common"
    HIGHLIGHT = "highlight"
    TITLE = "title"


# ======== DB 정보 DTO ========
@dataclass
class TableInfo:
    """
    테이블 정보

    Attributes:
        owner (str): 소유자, user
        table_name (str): 테이블명
        tablespace_name (Optional[str]): 테이블스페이스
        comment (Optional[str]): 테이블 설명명
    """

    owner: str
    table_name: str
    tablespace_name: Optional[str]
    comment: Optional[str]

    def __init__(self, table_info, dbms: RadioDbms):
        if dbms == RadioDbms.ORACLE:
            self.owner = table_info["OWNER"]
            self.table_name = table_info["TABLE_NAME"]
            self.tablespace_name = table_info["TABLESPACE_NAME"]
        elif dbms == RadioDbms.MYSQL:
            self.owner = table_info["TABLE_SCHEMA"].upper()
            self.table_name = table_info["TABLE_NAME"].upper()
            self.tablespace_name = None
            self.comment = table_info["TABLE_COMMENT"]


@dataclass
class ColumnInfo:
    """
    컬럼 정보

    Attributes:
        column_name (str): 컬럼명
        data_type (str): 컬럼타입
        data_length (Optional[int]): 데이터 길이
        nullable (Literal["Y", "N"]): NULL 허용 여부
        key_type (Optional[str]): PK, FK, IDX
        data_default (Optional[str]): 기본값
        comment (Optional[str]): 설명
    """

    column_name: str
    data_type: str
    data_length: Optional[int]
    nullable: Literal["Y", "N"]
    key_type: Optional[str]
    data_default: Optional[str]
    comment: Optional[str]

    def __init__(self, row, dbms: RadioDbms):
        if dbms == RadioDbms.ORACLE:
            self.column_name = row["COLUMN_NAME"]
            self.data_type = row["DATA_TYPE"]
            self.data_length = row["DATA_LENGTH"]
            self.nullable = row["NULLABLE"]
            self.key_type = row["KEY_TYPE"]
            self.data_default = row["DATA_DEFAULT"]
            self.comment = row["COL_COMMENTS"]
        elif dbms == RadioDbms.MYSQL:
            self.column_name = row["Field"].upper()
            self.data_type = row["Type"].upper()
            self.data_length = None
            self.nullable = "Y" if row["Null"] == "YES" else "N"
            self.key_type = row["Key"]
            self.data_default = row["Default"]
            self.comment = row["Comment"]


# ======== oracle client 경로 설정 ========
os.environ["PATH"] = Constant.LOCATION + ";" + os.environ["PATH"]

# ======== 전역변수 ========
# 완료 테이블 개수
TABLE_CNT = 0


def main():
    """
    진입점 함수
    """

    define_tk()
    draw_init_frame()

    TkObj.TK_ROOT.mainloop()


def define_tk():
    """
    GUI 정의
    """
    root = Tk()

    root.minsize(width=TkSize.WIDTH, height=TkSize.HEIGHT)
    root.title("테이블 명세서 생성")

    frame = Frame(root)
    frame.pack(padx=TkSize.GAP, pady=TkSize.GAP / 2)

    TkObj.TK_ROOT = root
    TkObj.TK_FRAME_FORM = frame


def draw_init_frame():
    """
    초기 frame 그리기
    """

    form_frame = TkObj.TK_FRAME_FORM
    form_frame.grid(row=0, column=0)

    # DBMS 종류 radio button
    Label(form_frame, text="DBMS").grid(row=0, column=0)
    radio_dbms_frame = Frame(form_frame)
    radio_dbms_frame.grid(row=0, column=1)
    radio_dbms = StringVar(value=RadioDbms.MYSQL.value)
    for d in RadioDbms:
        radio = Radiobutton(
            radio_dbms_frame,
            text=d.name,
            variable=radio_dbms,
            value=d.value,
        )
        radio.pack(side="left")

    # 스키마명
    Label(form_frame, text="스키마명").grid(row=1, column=0)
    schema_entry = Entry(form_frame)
    schema_entry.grid(row=1, column=1, ipadx=4, ipady=4)

    # 테이블명 리스트
    Label(form_frame, text="테이블명").grid(row=2, column=0)
    radio_table_list_frame = Frame(form_frame)
    radio_table_list_frame.grid(row=2, column=1)
    radio_table_list = StringVar(value=RadioTableList.ALL.name)
    for v in RadioTableList:
        radio = Radiobutton(
            radio_table_list_frame,
            text=v.value,
            variable=radio_table_list,
            value=v.name,
        )
        radio.pack(side="left")

    # 시트분리
    Label(form_frame, text="시트분리").grid(row=3, column=0)
    radio_sheet_frame = Frame(form_frame)
    radio_sheet_frame.grid(row=3, column=1)
    radio_sheet = StringVar(value=RadioSheet.MULTI.name)
    for v in RadioSheet:
        radio = Radiobutton(
            radio_sheet_frame, text=v.value, variable=radio_sheet, value=v.name
        )
        radio.pack(side="left")

    # 출력형식
    Label(form_frame, text="출력형식").grid(row=4, column=0)
    radio_output_type_frame = Frame(
        form_frame, width=TkSize.WIDTH, height=TkSize.HEIGHT / 2
    )
    radio_output_type_frame.grid(row=4, column=1)
    radio_output_type = StringVar(value=RadioOutputType.STYLED.name)
    for v in RadioOutputType:
        radio = Radiobutton(
            radio_output_type_frame,
            text=v.value,
            variable=radio_output_type,
            value=v.name,
        )
        radio.pack(side="left")

    # 출력경로
    Label(form_frame, text="파일경로\n(이름포함, 확장자 제외)").grid(
        row=5,
        column=0,
        padx=TkSize.GAP,
    )
    filename_entry = Entry(form_frame)
    filename_entry.grid(row=5, column=1, ipadx=4, ipady=4)

    Label(form_frame, text="작성자이름").grid(row=6, column=0, padx=TkSize.GAP)

    writer_entry = Entry(form_frame)
    writer_entry.grid(row=6, column=1, ipadx=4, ipady=4)

    run_btn = Button(form_frame, text="생성", command=run_btn_handler)
    run_btn.grid(
        row=7,
        column=0,
        columnspan=2,
        padx=TkSize.GAP,
        pady=TkSize.GAP,
    )

    TkObj.TK_FRAME_FORM = form_frame
    TkObj.TK_RADIO_DBMS = radio_dbms
    TkObj.TK_ENTRY_SCHEMA = schema_entry
    TkObj.TK_RADIO_TABLE_LIST = radio_table_list
    TkObj.TK_RADIO_SHEET = radio_sheet
    TkObj.TK_RADIO_OUTPUT_TYPE = radio_output_type
    TkObj.TK_ENTRY_OUTPUT_PATH = filename_entry


def run_btn_handler():
    """
    실행버튼 이벤트 핸들러
    """
    print("RUN!")

    # DBMS
    dbms: RadioDbms = RadioDbms[TkObj.TK_RADIO_DBMS.get()]
    # 스키마명
    schema = TkObj.TK_ENTRY_SCHEMA.get().upper()
    # 테이블명
    table_list_type = RadioTableList[TkObj.TK_RADIO_TABLE_LIST.get()]
    # 시트분리
    sheet = RadioSheet[TkObj.TK_RADIO_SHEET.get()]
    # 출력형식
    output_type = RadioOutputType[TkObj.TK_RADIO_OUTPUT_TYPE.get()]
    # 파일경로
    output_path = TkObj.TK_ENTRY_OUTPUT_PATH.get()
    # 작성자이름

    print(
        f"작성자명 {dbms}, 스키마명 {schema}, 테이블명 {table_list_type}, 시트분리 {sheet}, 출력형식 {output_type}, 출력경로 {output_path}"
    )

    # 실행
    exec(dbms, schema, table_list_type, sheet, output_type, output_path)


def define_excel_style(wb: Workbook):
    """
    엑셀 Named Style 정의
    """

    # 제목 셀
    title = NamedStyle(name=ExcelStyle.TITLE)
    title.fill = PatternFill(
        start_color="F2C9D1",
        end_color="F2C9D1",
        fill_type="solid",
    )
    title.alignment = Alignment(horizontal="center")
    title.font = Font(size=12, bold=True)

    # 강조 셀
    highlight = NamedStyle(name=ExcelStyle.HIGHLIGHT)
    highlight.fill = PatternFill(
        start_color="F2C9D1",
        end_color="F2C9D1",
        fill_type="solid",
    )
    highlight.alignment = Alignment(horizontal="center")
    highlight.font = Font(size=9, bold=True)

    # 전체 공통 적용 스타일(가운데정렬, 폰트사이즈)
    common = NamedStyle(name=ExcelStyle.COMMON)
    common.font = Font(size=9)
    common.alignment = Alignment(horizontal="center")

    wb.add_named_style(common)
    wb.add_named_style(highlight)
    wb.add_named_style(title)


def change_cells_style(row, style):
    """
    셀 스타일 변경
    """

    for c in row:
        c.style = style


def change_col_size(sheet: Worksheet):
    """
    시트 컬럼 사이즈 변경
    """

    ratio = 1.3
    sheet.column_dimensions["A"].width = 8 * ratio
    sheet.column_dimensions["B"].width = 30 * ratio
    sheet.column_dimensions["C"].width = 30 * ratio
    sheet.column_dimensions["D"].width = 12 * ratio
    sheet.column_dimensions["E"].width = 14 * ratio


def exec(
    dbms: RadioDbms,
    schema: str,
    table_list_type: RadioTableList,
    radio_sheet: RadioSheet,
    output_type: RadioOutputType,
    output_path: str,
):
    """
    실행
    """

    format = "%Y-%m-%d %H:%M:%S"

    print(f"----------- [{datetime.now().strftime(format)}] START TO SAVE EXCEL FILE")

    # 0. 엑셀 open
    wb = Workbook()

    # Named Style 정의
    define_excel_style(wb)

    try:
        conn = None

        # 1. DB 연결
        if dbms == RadioDbms.ORACLE:
            conn = cx_Oracle.connect(
                OracleConfig.USER,
                OracleConfig.PASSWORD,
                f"{OracleConfig.HOST}:{OracleConfig.PORT}/{OracleConfig.DATABASE}",
            )
        else:
            conn = pymysql.connect(
                host=MySqlConfig.HOST,
                port=MySqlConfig.PORT,
                user=MySqlConfig.USER,
                password=MySqlConfig.PASSWORD,
                database=MySqlConfig.DATABASE,
                charset=MySqlConfig.CHARSET,
                cursorclass=pymysql.cursors.DictCursor,
            )

        with conn.cursor() as cursor:

            if dbms == RadioDbms.ORACLE:
                exec_oracle(cursor)
            else:
                exec_mysql(
                    cursor,
                    wb,
                    schema,
                    table_list_type,
                    radio_sheet,
                    output_type,
                )

    except Exception as e:
        print(f"## FAIL TO SAVE FILE :: {e}")
        traceback.print_exc()
    else:
        print(f"SUCCESS SAVE EXCEL! :: {TABLE_CNT}")
    finally:
        wb.save(f"{output_path}.xlsx")

    print(f"----------- [{datetime.now().strftime(format)}] END TO SAVE EXCEL FILE")


def exec_mysql(
    cursor: pymysql.cursors.DictCursor,
    wb: Workbook,
    schema: str,
    table_list_type: RadioTableList,
    radio_sheet: RadioSheet,
    output_type: RadioOutputType,
):
    """
    MySQL 버전

    Args:
        cursor
        schema (str): DB Schema, 대문자
        table_list_type
        radio_sheet
        output_type
    """

    global TABLE_CNT

    # excel 시트
    sheet: Worksheet = wb.active if radio_sheet == RadioSheet.ONE else None

    # 2. 대상 테이블명 리스트 조회
    tables = None

    if table_list_type == RadioTableList.ALL:
        # 전체 테이블
        sql = f"SHOW TABLES FROM {schema} WHERE Tables_in_{schema.lower()}"
        cursor.execute(sql)
        tables = cursor.fetchall()
    else:
        # 특정 테이블
        pass

    # 3. 테이블정보 조회
    for t in tables:
        table_name: str = t[f"Tables_in_{schema.lower()}"]
        sql = f"""
            SELECT * FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_NAME = '{table_name}'
            """
        cursor.execute(sql)
        table_info = cursor.fetchone()

        # 4. 컬럼정보 조회
        sql = f"SHOW FULL COLUMNS FROM {table_name}"
        cursor.execute(sql)
        column_list = cursor.fetchall()

        # TODO - 시트 셋팅
        if radio_sheet == RadioSheet.MULTI:
            pass

        # 5. 엑셀 write
        write_excel(
            sheet,
            table_info=TableInfo(table_info, RadioDbms.MYSQL),
            column_list=column_list,
            dbms=RadioDbms.MYSQL,
            output_type=output_type,
        )

        # 5-1. 시트 컬럼사이즈 변경
        change_col_size(sheet)

        # TODO - 6. 진행 완료 테이블수 화면에 업데이트
        TABLE_CNT += 1


def exec_oracle(cursor):
    """
    Oracle
    """
    # 2. 대상 테이블명 리스트 조회

    # 3. 테이블정보 조회

    # 4. 컬럼정보 조회

    # 5. 엑셀 write


def write_excel(
    sheet: Worksheet,
    table_info: TableInfo,
    column_list: List[dict],
    dbms: RadioDbms,
    output_type: RadioOutputType,
):
    """
    엑셀에 작성
    """

    sheet.append(["테이블정의서"])
    max_row = sheet.max_row
    sheet.merge_cells(f"A{max_row}:H{max_row}")
    output_type == RadioOutputType.STYLED and change_cells_style(
        sheet[f"{max_row}:{max_row}"], ExcelStyle.TITLE
    )

    # 테이블 정보
    sheet.append(
        [
            "테이블명",
            table_info.table_name,
            "",
            "",
            "테이블 스페이스",
            table_info.tablespace_name,
            "",
            "",
        ]
    )
    max_row = sheet.max_row
    sheet.merge_cells(f"B{max_row}:D{max_row}")
    sheet.merge_cells(f"F{max_row}:H{max_row}")
    change_cells_style(sheet[f"{max_row}:{max_row}"], ExcelStyle.COMMON)
    if output_type == RadioOutputType.STYLED:
        change_cells_style(
            [
                sheet[f"A{max_row}"],
                sheet[f"E{max_row}"],
                sheet[f"H{max_row}"],
            ],
            ExcelStyle.HIGHLIGHT,
        )
        # TODO - 왼쪽정렬

    sheet.append(["테이블 설명", table_info.comment, "", "", "", "", "", ""])
    max_row = sheet.max_row
    sheet.merge_cells(f"B{max_row}:H{max_row}")
    change_cells_style(sheet[f"{max_row}:{max_row}"], ExcelStyle.COMMON)
    output_type == RadioOutputType.STYLED and change_cells_style(
        [sheet[f"A{max_row}"], sheet[f"E{max_row}"]], ExcelStyle.HIGHLIGHT
    )

    # 컬럼 정보
    sheet.append(["번호", "컬럼명", "설명", "타입", "길이", "NULL", "KEY", "기본값"])
    max_row = sheet.max_row
    output_type == RadioOutputType.STYLED and change_cells_style(
        sheet[f"{max_row}:{max_row}"], ExcelStyle.HIGHLIGHT
    )

    i = 0
    for c in column_list:
        i += 1
        column_info = ColumnInfo(c, dbms)
        sheet.append(
            [
                i,
                column_info.column_name,
                column_info.comment,
                column_info.data_type,
                column_info.data_length,
                column_info.nullable,
                column_info.key_type,
                column_info.data_default,
            ]
        )

        max_row = sheet.max_row
        change_cells_style(sheet[f"{max_row}:{max_row}"], ExcelStyle.COMMON)


main()
